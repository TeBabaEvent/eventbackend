#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export paid tickets for one event into an Excel file.

Examples:
  python export_event_tickets.py --today
  python export_event_tickets.py --event-id 123e4567-e89b-12d3-a456-426614174000
  python export_event_tickets.py --event-slug baba-night
  python export_event_tickets.py --event-date 2026-02-14 --event-title "BABA Night"
  python export_event_tickets.py --today --output exports/tickets-today.xlsx
"""

import argparse
import io
import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal
from typing import List, Sequence, Tuple

from sqlalchemy.orm import joinedload

from app.db.database import SessionLocal
from app.db import models

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore[assignment]


# Fix encoding for Windows console
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


def parse_iso_date(value: str) -> str:
    """Validate YYYY-MM-DD date format."""
    try:
        return datetime.strptime(value, "%Y-%m-%d").date().isoformat()
    except ValueError as exc:
        raise argparse.ArgumentTypeError(
            f"Invalid date '{value}'. Expected format: YYYY-MM-DD."
        ) from exc


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser(
        description="Export paid tickets (completed orders only) for one event into Excel."
    )
    parser.add_argument("--today", action="store_true", help="Shortcut for --event-date set to today.")
    parser.add_argument("--event-id", help="Event UUID.")
    parser.add_argument("--event-slug", help="Event slug.")
    parser.add_argument("--event-date", type=parse_iso_date, help="Event date (YYYY-MM-DD).")
    parser.add_argument("--event-title", help="Partial event title filter.")
    parser.add_argument("--output", help="Output .xlsx path (default: exports/tickets-<event>-<date>.xlsx).")

    args = parser.parse_args()

    if args.today and args.event_date:
        parser.error("Use either --today or --event-date, not both.")

    return args


def normalize_filename(value: str, fallback: str) -> str:
    """Return a filesystem-safe token."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip())
    cleaned = cleaned.strip("-._")
    return cleaned or fallback


def resolve_output_path(event: models.Event, custom_output: str | None) -> str:
    """Build output path and ensure parent folder exists."""
    if custom_output:
        output_path = custom_output
        if not output_path.lower().endswith(".xlsx"):
            output_path = f"{output_path}.xlsx"
        parent = os.path.dirname(os.path.abspath(output_path))
        if parent:
            os.makedirs(parent, exist_ok=True)
        return os.path.abspath(output_path)

    export_dir = "exports"
    os.makedirs(export_dir, exist_ok=True)

    event_token = normalize_filename(event.slug or event.title or event.id, event.id)
    filename = f"tickets-{event_token}-{event.date}.xlsx"
    output_path = os.path.join(export_dir, filename)

    if os.path.exists(output_path):
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        output_path = os.path.join(export_dir, f"tickets-{event_token}-{event.date}-{stamp}.xlsx")

    return output_path


def get_matching_events(db, args: argparse.Namespace, target_date: str | None) -> List[models.Event]:
    """Return events matching filters."""
    query = db.query(models.Event)

    if args.event_id:
        query = query.filter(models.Event.id == args.event_id)
    if args.event_slug:
        query = query.filter(models.Event.slug == args.event_slug)
    if target_date:
        query = query.filter(models.Event.date == target_date)
    if args.event_title:
        query = query.filter(models.Event.title.ilike(f"%{args.event_title}%"))

    return query.order_by(models.Event.date.asc(), models.Event.time.asc()).all()


def get_paid_orders_for_event(db, event_id: str) -> List[models.Order]:
    """Fetch paid orders only (status=completed)."""
    return (
        db.query(models.Order)
        .options(
            joinedload(models.Order.items).joinedload(models.OrderItem.pack),
            joinedload(models.Order.pack),
            joinedload(models.Order.tickets),
        )
        .filter(
            models.Order.event_id == event_id,
            models.Order.status == "completed",
        )
        .order_by(models.Order.paid_at.asc(), models.Order.created_at.asc())
        .all()
    )


def format_datetime(value: datetime | None) -> str:
    """Format datetime for export."""
    if not value:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def build_rows(orders: Sequence[models.Order]) -> Tuple[List[Tuple], int, Decimal]:
    """Build export rows and totals."""
    rows: List[Tuple] = []
    total_tickets = 0
    total_amount = Decimal("0.00")

    for order in orders:
        tickets_count = int(order.total_quantity or 0)
        amount_eur = (Decimal(order.amount or 0) / Decimal("100")).quantize(Decimal("0.01"))
        payment_method = getattr(order, "payment_method", "online") or "online"

        rows.append(
            (
                order.order_number,
                order.customer_name,
                order.customer_email,
                order.customer_phone or "",
                tickets_count,
                float(amount_eur),
                payment_method,
                format_datetime(order.paid_at),
                order.pack_display,
            )
        )

        total_tickets += tickets_count
        total_amount += amount_eur

    return rows, total_tickets, total_amount


def autosize_columns(ws) -> None:
    """Adjust column width for readability."""
    for column_index in range(1, ws.max_column + 1):
        col_letter = get_column_letter(column_index)
        max_len = 0

        for row_index in range(1, ws.max_row + 1):
            value = ws.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))

        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))


def write_excel(
    output_path: str,
    event: models.Event,
    rows: Sequence[Tuple],
    total_orders: int,
    total_tickets: int,
    total_amount: Decimal,
) -> None:
    """Generate XLSX file."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tickets payes"

    title_font = Font(size=14, bold=True)
    bold_font = Font(bold=True)
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    center_align = Alignment(horizontal="center", vertical="center")

    sheet["A1"] = "Export tickets payes"
    sheet["A1"].font = title_font
    sheet["A2"] = "Evenement"
    sheet["B2"] = event.title
    sheet["A3"] = "Date evenement"
    sheet["B3"] = event.date
    sheet["A4"] = "Genere le"
    sheet["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet["A5"] = "Statut des commandes incluses"
    sheet["B5"] = "completed (payees)"

    header_row = 7
    headers = [
        "Commande",
        "Client",
        "Email",
        "Telephone",
        "Tickets",
        "Montant paye (EUR)",
        "Methode paiement",
        "Paye le",
        "Packs",
    ]

    for col_index, label in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_offset, row_values in enumerate(rows, start=1):
        row_index = header_row + row_offset
        for col_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            if col_index == 5:
                cell.alignment = center_align
            if col_index == 6:
                cell.number_format = "#,##0.00"

    total_row = header_row + len(rows) + 2
    sheet.cell(row=total_row, column=1, value="TOTAL").font = bold_font
    sheet.cell(row=total_row, column=2, value=f"{total_orders} commandes payees").font = bold_font
    sheet.cell(row=total_row, column=5, value=total_tickets).font = bold_font
    total_amount_cell = sheet.cell(row=total_row, column=6, value=float(total_amount))
    total_amount_cell.font = bold_font
    total_amount_cell.number_format = "#,##0.00"

    last_table_row = max(header_row, header_row + len(rows))
    sheet.auto_filter.ref = f"A{header_row}:I{last_table_row}"
    sheet.freeze_panes = f"A{header_row + 1}"

    autosize_columns(sheet)
    workbook.save(output_path)


def main() -> int:
    """CLI entrypoint."""
    args = parse_args()

    if Workbook is None:
        print("Error: openpyxl is missing. Run: pip install -r requirements.txt")
        return 1

    if not any([args.today, args.event_id, args.event_slug, args.event_date, args.event_title]):
        args.today = True
        print(f"No event filter provided. Using today's date: {date.today().isoformat()}")

    target_date = date.today().isoformat() if args.today else args.event_date

    db = SessionLocal()
    try:
        events = get_matching_events(db, args, target_date)

        if not events:
            print("No event found with these filters.")
            if target_date:
                print(f"Date filter: {target_date}")
            return 1

        if len(events) > 1:
            print("Multiple events found. Refine your filters with --event-id or --event-slug:")
            for event in events:
                print(f"- {event.id} | {event.date} {event.time} | {event.title}")
            return 1

        event = events[0]
        orders = get_paid_orders_for_event(db, event.id)
        rows, total_tickets, total_amount = build_rows(orders)

        output_path = resolve_output_path(event, args.output)
        write_excel(
            output_path=output_path,
            event=event,
            rows=rows,
            total_orders=len(orders),
            total_tickets=total_tickets,
            total_amount=total_amount,
        )

        print("Export generated successfully.")
        print(f"Event: {event.title} ({event.date})")
        print(f"Paid orders: {len(orders)}")
        print(f"Tickets: {total_tickets}")
        print(f"Revenue (EUR): {float(total_amount):.2f}")
        print(f"File: {output_path}")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
